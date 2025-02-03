import os
import shutil
from openpyxl import load_workbook
from oletools.olevba3 import VBA_Parser

EXCEL_FILE_EXTENSIONS = ("xlsb", "xls", "xlsm", "xla", "xlt", "xlam", "xlsx")
KEEP_NAME = False  # Set this to True if you want to keep "Attribute VB_Name"


def parse_vba(workbook_path, file_prefix):
    """Extracts VBA macros from Excel files and saves them with a file prefix."""
    vba_path = "src.vba"
    vba_parser = VBA_Parser(workbook_path)
    vba_modules = (
        vba_parser.extract_all_macros() if vba_parser.detect_vba_macros() else []
    )

    for _, _, filename, content in vba_modules:
        lines = content.splitlines()
        content_filtered = [
            line
            for line in lines
            if not line.startswith("Attribute") or ("VB_Name" in line and KEEP_NAME)
        ]

        # Save VBA macro with a unique name per file
        if content_filtered:
            os.makedirs(vba_path, exist_ok=True)
            with open(
                os.path.join(vba_path, f"{file_prefix}_{filename}"),
                "w",
                encoding="utf-8",
            ) as f:
                f.write("\n".join(content_filtered))


def generate_text_reports(workbook_path, file_prefix):
    """Generates text reports for all details of an Excel file, including named ranges."""
    report_path = "excel_reports"
    os.makedirs(report_path, exist_ok=True)

    workbook = load_workbook(workbook_path, data_only=False)  # Formulas
    eval_workbook = load_workbook(workbook_path, data_only=True)  # Evaluated values

    reports = {
        f"{file_prefix}_formulas_and_values.txt": [],
        f"{file_prefix}_formatting.txt": [],
        f"{file_prefix}_conditional_formatting.txt": [],
        f"{file_prefix}_merged_cells.txt": [],
        f"{file_prefix}_data_validations.txt": [],
        f"{file_prefix}_hyperlinks.txt": [],
        f"{file_prefix}_named_ranges.txt": [],
    }

    # Track Named Ranges
    named_ranges_section = ["Workbook Named Ranges\n" + "-" * 40]
    for name in workbook.defined_names:  # ✅ Corrected iteration
        destination = workbook.defined_names[name].attr_text  # Named range reference
        named_ranges_section.append(f"Name: {name}, Refers To: {destination}")
    reports[f"{file_prefix}_named_ranges.txt"].extend(named_ranges_section)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        eval_sheet = eval_workbook[sheet_name]

        # 1. Formulas & Values
        formulas_section = [f"Sheet: {sheet_name}\n" + "-" * 40]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Determine the type of the cell content
                    if isinstance(cell.value, str):
                        cell_type = "Text"
                    elif isinstance(cell.value, (int, float)):
                        cell_type = "Number"
                    elif isinstance(cell.value, bool):
                        cell_type = "Boolean"
                    else:
                        cell_type = "Other"

                    # Format the output to include the type and value
                    formulas_section.append(
                        f"Cell {cell.coordinate}: Value='{cell.value}', Type='{cell_type}'"
                    )
        reports[f"{file_prefix}_formulas_and_values.txt"].extend(formulas_section)

        # 2. Formatting
        formatting_section = [f"Sheet: {sheet_name}\n" + "-" * 40]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    formatting_section.append(
                        f"Cell {cell.coordinate}: Font='{cell.font.name}', Size={cell.font.size}, "
                        f"Bold={cell.font.bold}, Italic={cell.font.italic}, "
                        f"Font Color={cell.font.color.rgb if cell.font.color else 'None'}, "
                        f"Fill Color={cell.fill.start_color.rgb if cell.fill.start_color else 'None'}, "
                        f"Alignment={cell.alignment.horizontal}, Number Format='{cell.number_format}'"
                    )
        reports[f"{file_prefix}_formatting.txt"].extend(formatting_section)

        # 3. Conditional Formatting
        conditional_section = [f"Sheet: {sheet_name}\n" + "-" * 40]
        if sheet.conditional_formatting:
            for rule in sheet.conditional_formatting:
                conditional_section.append(
                    f"Rule: {rule}, Applied to: {sheet.conditional_formatting[rule]}"
                )
        reports[f"{file_prefix}_conditional_formatting.txt"].extend(conditional_section)

        # 4. Merged Cells
        merged_section = [f"Sheet: {sheet_name}\n" + "-" * 40]
        for merged_range in sheet.merged_cells.ranges:
            merged_section.append(f"Merged Range: {merged_range}")
        reports[f"{file_prefix}_merged_cells.txt"].extend(merged_section)

        # 5. Data Validations
        validations_section = [f"Sheet: {sheet_name}\n" + "-" * 40]
        if sheet.data_validations:
            for dv in sheet.data_validations.dataValidation:
                validations_section.append(
                    f"Range: {dv.sqref}, Formula: {dv.formula1}, Allow Type: {dv.type}, Criteria: {dv.operator}"
                )
        reports[f"{file_prefix}_data_validations.txt"].extend(validations_section)

        # 6. Hyperlinks
        hyperlinks_section = [f"Sheet: {sheet_name}\n" + "-" * 40]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.hyperlink:
                    hyperlinks_section.append(
                        f"Cell {cell.coordinate}: Hyperlink='{cell.hyperlink.target}'"
                    )
        reports[f"{file_prefix}_hyperlinks.txt"].extend(hyperlinks_section)

    # Save reports to text files
    for filename, content in reports.items():
        with open(os.path.join(report_path, filename), "w", encoding="utf-8") as f:
            f.write("\n".join(content) + "\n")

    print(f"Reports for {file_prefix} saved in '{report_path}' folder.")


if __name__ == "__main__":
    # Remove old VBA and Excel reports before processing
    if os.path.exists("src.vba"):
        shutil.rmtree("src.vba")
    if os.path.exists("excel_reports"):
        shutil.rmtree("excel_reports")

    # Scan current directory for Excel files
    for root, _, files in os.walk("."):
        for file in files:
            if file.endswith(EXCEL_FILE_EXTENSIONS):
                file_path = os.path.join(root, file)
                file_prefix = os.path.splitext(file)[
                    0
                ]  # Get filename without extension
                print(f"Processing: {file_path}")

                # Extract VBA macros
                parse_vba(file_path, file_prefix)

                # Generate reports
                generate_text_reports(file_path, file_prefix)
